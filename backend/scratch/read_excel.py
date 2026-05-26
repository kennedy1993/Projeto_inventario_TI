import openpyxl

wb = openpyxl.load_workbook("Inventário de Ativos de T.I.xlsx")
print("Sheets:", wb.sheetnames)

for name in wb.sheetnames:
    sheet = wb[name]
    print(f"\nSheet: {name}")
    # Print first few rows that contain data
    rows = list(sheet.iter_rows(values_only=True))
    for i, r in enumerate(rows[:5]):
        if any(r):
            print(f"  Row {i+1}:", r[:10])

